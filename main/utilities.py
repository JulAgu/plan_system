import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill
from itertools import combinations
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

CONFIG_NAME = "config.xlsx"

class Microparcelle(object):
    """
    Un objet gardat tous les attributs d"une microparcelle
    """

    def __init__(self, nom, position, traitement):
        self.nom = nom
        self.position = position
        self.traitement = traitement

    def __str__(self):
        return "Nom: {0}, Position: {1}, Traitement: {2} ".format(
            self.nom, self.position, self.traitement
        )


class Essai(object):
    """
    Comporte par des microparcelles, l'essai garde les attributs et la hierarchie des microparcelles
    """

    def __init__(self, nom, position, pathData, configPath):
        self.nom = nom
        self.position = position
        self.pathData = pathData
        self.configPath = configPath

    def __str__(self):
        return "Nom: {0}, Position: {1}, Path: {2} ".format(
            self.nom, self.position, self.pathData
        )

    def __len__(self):
        try:
            return len(self.listDeMicroParcelles)
        except:
            return 0

    def createEssai(self):
        data = pd.read_excel(self.pathData, sheet_name="Données_1")
        corr = pd.read_excel(self.pathData, sheet_name="Correspondance")
        config = pd.read_excel(self.configPath)
        config.drop_duplicates(subset="Variété", keep="first", inplace=True)
        self.descriptionData = corr.loc[:, ["Nom Variété", "Numéro de correspondance"]]
        self.descriptionData = self.descriptionData.rename(
            columns={"Nom Variété": "VARIETES", "Numéro de correspondance": "N"}
        )
        self.descriptionData["PROTOCOLE"] = self.nom
        self.descriptionData = pd.merge(
            self.descriptionData,
            config,
            left_on="VARIETES",
            right_on="Variété",
            how="left",
        )
        self.descriptionData["P.SAC"] = ""
        self.descriptionData["DENSITE"] = ""
        self.descriptionData.rename(
            columns={"Semencier": "FIRME", "Espèce": "ESPECE"}, inplace=True
        )
        self.descriptionData = self.descriptionData.loc[
            :,
            [
                "VARIETES",
                "ESPECE",
                "N",
                "PROTOCOLE",
                "FIRME",
                "P.SAC",
                "PMG",
                "DENSITE",
            ],
        ]
        self.descriptionData = self.descriptionData.fillna("")
        self.descriptionData = self.descriptionData.drop_duplicates()

        correspondance = {
            corr["Plan"][i]: corr["Numéro de correspondance"][i]
            for i in range(len(corr))
        }
        data["N° de traitement"] = data["N° de traitement"].apply(
            lambda x: correspondance[x]
        )
        data = data.sort_values(by="Position")
        self.nModalite = data["N° de traitement"].nunique()
        self.nBlocs = (data["Position"] // 100).nunique()

        data = data.apply(lineToMicroparcelle, axis=1)
        self.listDeMicroParcelles = list(data)

    def create_matrice(self):
        """
        Crée une matrice ordonnée de microparcelles
        """
        self.matrice = np.zeros((self.nModalite, self.nBlocs))
        nomMicroParcelles = [x.nom for x in self.listDeMicroParcelles]
        for i in range(self.nBlocs):
            self.matrice[:, i] = nomMicroParcelles[
                i * self.nModalite : (i + 1) * self.nModalite
            ]


class Dessinateur4blocs(object):
    """
    Classe contenant tous les attributs et méthodes nécessaires pour dessiner
    les objets Essai dans un seul fichier Excel.
    """

    def __init__(self, nomDuPlan, essais, essaisMatricesList, droiteIndex, gaucheIndex):
        self.nomDuPlan = "/".join(nomDuPlan.split("/")[:-1]) + "/results/" + nomDuPlan.split("/")[-1] 
        self.nomAffichage = nomDuPlan.split("/")[-1]
        self.essais = essais
        self.essaisMatricesList = essaisMatricesList
        self.droiteIndex = droiteIndex
        self.gaucheIndex = gaucheIndex
        self.initialRow = 2
        self.initialCol = 0
        self.ppInitialRow = 2
        self.ppInitialCol = 9
        self.colors = [
            "FDADD4",
            "b3cde3",
            "ccebc5",
            "8DC4F7",
            "fed9a6",
            "ffffcc",
            "e5d8bd",
            "A2D7F5",
            "decbe4",
        ]

    def openDraw(self):
        self.workbook = xlsxwriter.Workbook(self.nomDuPlan + ".xlsx")
        self.worksheet = self.workbook.add_worksheet("Plan")
        self.simpleCellFormat = self.workbook.add_format({"border": 1})
        self.centerCellFormat = self.workbook.add_format(
            {"align": "center", "border": 1}
        )
        self.countCellFormat = self.workbook.add_format({"bold": True, "border": 1})

    def closeDraw(self):
        self.workbook.close()

    def planPrincipal(self):
        row = self.ppInitialRow
        col = self.ppInitialCol

        leftMatriceList = [self.essaisMatricesList[i] for i in self.gaucheIndex]
        rightMatriceList = [self.essaisMatricesList[i] for i in self.droiteIndex]

        for matrice in leftMatriceList:
            matrice = pd.DataFrame(matrice)
            for bloc in matrice.columns:
                self.worksheet.write_column(
                    row, col, list(matrice[bloc]), self.centerCellFormat
                )
                col += 2
            row += len(matrice)
            col = self.ppInitialCol

        row = self.ppInitialRow
        col = self.ppInitialCol + leftMatriceList[0].shape[1] * 2

        for matrice in rightMatriceList:
            matrice = pd.DataFrame(matrice)
            matrice = matrice[matrice.columns[::-1]]
            for bloc in matrice.columns:
                self.worksheet.write_column(
                    row, col, list(matrice[bloc]), self.centerCellFormat
                )
                col += 2
            row += len(matrice)
            col = self.ppInitialCol + leftMatriceList[0].shape[1] * 2

    def drawBordures(self):
        row = self.initialRow
        col = self.initialCol

        self.essaisGauche = [self.essais[i] for i in self.gaucheIndex]
        self.essaisDroite = [self.essais[i] for i in self.droiteIndex]

        self.dfGauche = pd.concat(
            [essai.descriptionData for essai in self.essaisGauche]
        ).reset_index(drop=True)
        self.dfDroite = pd.concat(
            [essai.descriptionData for essai in self.essaisDroite]
        ).reset_index(drop=True)

        for colone in self.dfGauche.columns:
            self.worksheet.write(
                row - 1, col, colone, self.workbook.add_format({"bold": True})
            )
            self.worksheet.write_column(
                row, col, list(self.dfGauche[colone]), self.simpleCellFormat
            )
            col += 1

        col = (
            (self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4) + 1 + 1
        )  # +1 pour laisser un espace et +1 pour la colonne d'énumeration

        for colone in self.dfDroite.columns:
            self.worksheet.write(
                row - 1, col, colone, self.workbook.add_format({"bold": True})
            )
            self.worksheet.write_column(
                row, col, list(self.dfDroite[colone]), self.simpleCellFormat
            )
            col += 1

    def matriceCompteur(self):
        """
        Compte les microparcelles de chaque essai
        """
        matriceComptage = np.zeros(
            (
                sum(essai.nModalite for essai in self.essaisGauche),
                self.essaisGauche[0].nBlocs + self.essaisDroite[0].nBlocs,
            )
        )

        matriceComptage[0, :] = range(1, matriceComptage.shape[1] + 1)
        for i in range(len(matriceComptage) - 1):
            matriceComptage[i + 1, :] = range(
                int(matriceComptage[i, -1] + 1),
                int(matriceComptage[i, -1]) + matriceComptage.shape[1] + 1,
            )

        for i, row in enumerate(matriceComptage):
            if i % 2 == 0:
                matriceComptage[i] = row[::-1]

        row = self.ppInitialRow
        col = self.ppInitialCol + 1

        self.matriceComptage = pd.DataFrame(matriceComptage)
        for colonne in self.matriceComptage.columns:
            self.worksheet.write_column(
                row, col, list(self.matriceComptage[colonne]), self.countCellFormat
            )
            col += 2

    def createPavePourLesCalculs(self):
        """
        Creer un pave pour les calculs pour les poids des sacs
        en fonction de la surface de chaque microparcelle
        """
        worksheetPave = self.workbook.add_worksheet("Calcul")

        worksheetPave.merge_range("B2:C2", "Dimensions Parcelle", self.countCellFormat)
        worksheetPave.write("B3", "Longueur", self.countCellFormat)
        worksheetPave.write("B4", "Largeur", self.countCellFormat)
        worksheetPave.write("C3", 9, self.simpleCellFormat)
        worksheetPave.write("C4", 1.36, self.simpleCellFormat)

        worksheetPave.merge_range("E2:F2", "Densité de semis", self.countCellFormat)
        worksheetPave.write("E3", "Espèce", self.countCellFormat)
        worksheetPave.write("F3", "Densité", self.countCellFormat)

        for rowCounter, especeName, densiteEspece in zip(
            [i for i in range(4, 11)],
            [
                j
                for j in [
                    "Blé lignée",
                    "Blé hybride",
                    "Triticale",
                    "Orge 6R",
                    "Orge 2R",
                    "Orge hybride",
                    "Orge de Printemps",
                ]
            ],
            [280, 160, 180, 230, 260, 180, 30],
        ):

            worksheetPave.write(f"E{rowCounter}", especeName, self.simpleCellFormat)
            worksheetPave.write(f"F{rowCounter}", densiteEspece, self.simpleCellFormat)

        worksheetPave.autofit()

    def fisher(self):
        """
        Integre les formules mathematiques pour le calcul des poids des sacs
        """
        lenConfig = len(pd.read_excel(CONFIG_NAME))+1

        for i in range(len(self.dfGauche)):
            self.worksheet.write_formula(
                f"F{3+i}",
                f"=G{3+i}*H{3+i}*Calcul!$C$3*Calcul!$C$4/1000",
                self.simpleCellFormat,
            )

        for i in range(len(self.dfDroite)):
            self.worksheet.write_formula(
                f"AG{3+i}",
                f"=AH{3+i}*AI{3+i}*Calcul!$C$3*Calcul!$C$4/1000",
                self.simpleCellFormat,
            )

        # LookUp pour les densités grace au pave de la feuille Calcul
        for i in range(3, len(self.dfGauche) + 3):
            self.worksheet.write_formula(
                f"H{i}",
                f"=XLOOKUP(B{i},Calcul!E4:E10,Calcul!F4:F10,0)",
                self.simpleCellFormat,
            )

        for i in range(3, len(self.dfDroite) + 3):
            self.worksheet.write_formula(
                f"AI{i}",
                f"=XLOOKUP(AC{i},Calcul!E4:E10,Calcul!F4:F10,0)",
                self.simpleCellFormat,
            )
        # LookUp pour les PMGs qui s'actualisent en temps reél
        for i in range(3, len(self.dfGauche) + 3):
            self.worksheet.write_formula(
                f"G{i}",
                f"=XLOOKUP(A{i},'{os.getcwd()}\[config.xlsx]Feuil1'!B2:B{lenConfig},'{os.getcwd()}\[config.xlsx]Feuil1'!D2:D{lenConfig},0)",
                self.simpleCellFormat,
            )
        
        for i in range(3, len(self.dfGauche) + 3):
            self.worksheet.write_formula(
                f"AH{i}",
                f"=XLOOKUP(AB{i},'{os.getcwd()}\[config.xlsx]Feuil1'!B2:B{lenConfig},'{os.getcwd()}\[config.xlsx]Feuil1'!D2:D{lenConfig},0)",
                self.simpleCellFormat,
            )

    def etiquettes(self):
        """
        Construit l'onglet dediée à l'impression des etiquettes
        """

        matriceComptageGauche = self.matriceComptage.iloc[
            :, : self.essaisGauche[0].nBlocs
        ]
        matriceComptageDroite = self.matriceComptage.iloc[
            :, self.essaisGauche[0].nBlocs :
        ]

        leftMatriceList = [self.essaisMatricesList[i] for i in self.gaucheIndex]
        rightMatriceList = [self.essaisMatricesList[i] for i in self.droiteIndex]

        startPoint = 0
        etiquettesGauche = pd.DataFrame(columns=["Position", "Parcelle", "Essai"])
        for essai, matrice in zip(self.essaisGauche, leftMatriceList):
            position = (
                matriceComptageGauche[startPoint : startPoint + essai.nModalite]
                .to_numpy()
                .flatten()
            )
            parcelle = matrice.flatten()
            newEtiquetets = pd.DataFrame(
                {"Position": position, "Parcelle": parcelle, "Essai": essai.nom}
            )
            newEtiquetets["Modalité"] = (
                newEtiquetets["Parcelle"] - newEtiquetets["Parcelle"] // 100 * 100
            )
            newEtiquetets = pd.merge(
                newEtiquetets,
                essai.descriptionData.loc[:, ["VARIETES", "N"]],
                left_on="Modalité",
                right_on="N",
            )
            etiquettesGauche = pd.concat([etiquettesGauche, newEtiquetets])
            startPoint += essai.nModalite

        startPoint = 0
        etiquettesDroite = pd.DataFrame(
            columns=["Position", "Parcelle", "Essai", "Modalité"]
        )
        for essai, matrice in zip(self.essaisDroite, rightMatriceList):
            position = (
                matriceComptageDroite[startPoint : startPoint + essai.nModalite]
                .to_numpy()
                .flatten()
            )
            matrice = pd.DataFrame(matrice)
            matrice = matrice[matrice.columns[::-1]]
            parcelle = matrice.to_numpy().flatten()
            newEtiquetets = pd.DataFrame(
                {"Position": position, "Parcelle": parcelle, "Essai": essai.nom}
            )
            newEtiquetets["Modalité"] = (
                newEtiquetets["Parcelle"] - newEtiquetets["Parcelle"] // 100 * 100
            )
            newEtiquetets = pd.merge(
                newEtiquetets,
                essai.descriptionData.loc[:, ["VARIETES", "N"]],
                left_on="Modalité",
                right_on="N",
            )
            etiquettesDroite = pd.concat([etiquettesDroite, newEtiquetets])
            startPoint += essai.nModalite

        etiquettes = pd.concat([etiquettesGauche, etiquettesDroite])
        etiquettes = etiquettes.loc[
            :, ["Position", "Parcelle", "Essai", "VARIETES", "Modalité"]
        ]
        etiquettes["Plan"] = self.nomAffichage
        etiquettes = etiquettes.rename(columns={"VARIETES": "Variété"})
        etiquettes = etiquettes.sort_values(by="Position")
        etiquettes = etiquettes.reset_index(drop=True)

        worksheetEtiquettes = self.workbook.add_worksheet("Etiquettes")
        for colonne in etiquettes.columns:
            worksheetEtiquettes.write_column(
                1,
                list(etiquettes.columns).index(colonne),
                list(etiquettes[colonne]),
                self.simpleCellFormat,
            )
        worksheetEtiquettes.write_row(
            0, 0, list(etiquettes.columns), self.countCellFormat
        )

        worksheetEtiquettes.write(
            0, len(etiquettes.columns), "P.SAC", self.countCellFormat
        )

        for i in range(2, len(etiquettes) + 2):
            worksheetEtiquettes.write_formula(
                f"H{i}",
                f"=XLOOKUP(C{i}&D{i},Plan!AE{self.ppInitialRow+1}:AE{self.ppInitialRow + len(self.dfDroite)}&Plan!AB{self.ppInitialRow+1}:AB{self.ppInitialRow + len(self.dfDroite)},Plan!AG{self.ppInitialRow+1}:AG{self.ppInitialRow + len(self.dfDroite)},0)",
            )

            worksheetEtiquettes.write_formula(
                f"G{i}",
                f"=XLOOKUP(C{i}&D{i},Plan!D{self.ppInitialRow+1}:D{self.ppInitialRow + len(self.dfGauche)}&Plan!A{self.ppInitialRow+1}:A{self.ppInitialRow + len(self.dfGauche)},Plan!F{self.ppInitialRow+1}:F{self.ppInitialRow + len(self.dfGauche)},H{i})",
                self.simpleCellFormat,
            )
        worksheetEtiquettes.autofit()

    def maquillage(self):
        """
        Regle les proportions des celules et effectue la mise en page finale
        """
        self.worksheet.autofit()
        # Regler les proportions des colonnes
        for i in range(
            self.ppInitialCol,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4,
            2,
        ):
            self.worksheet.set_column(i, i, 12)

        # Mettre le titre dans le plan principal
        self.worksheet.merge_range(
            self.ppInitialRow - 2,
            self.ppInitialCol,
            self.ppInitialRow - 2,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4 - 1,
            self.nomAffichage,
            self.workbook.add_format({"bold": "true", "align": "center", "border": 1}),
        )

        # Fusionner les cellules des bordures
        for i in range(
            self.ppInitialCol,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4,
            2,
        ):
            self.worksheet.merge_range(
                self.ppInitialRow - 1,
                i,
                self.ppInitialRow - 1,
                i + 1,
                "",
                self.workbook.add_format(
                    {"bold": "true", "align": "center", "border": 1}
                ),
            )
            self.worksheet.merge_range(
                self.ppInitialRow + len(self.dfGauche),
                i,
                self.ppInitialRow + len(self.dfGauche),
                i + 1,
                "",
                self.workbook.add_format(
                    {"bold": "true", "align": "center", "border": 1}
                ),
            )

        # Add the counter at right of the plan
        col = self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4
        row = self.ppInitialRow - 1
        for index, i in enumerate(range(row, row + len(self.dfGauche) + 2)):
            self.worksheet.write(
                i, col,
                index + 1,
                self.workbook.add_format({"align": "left"})
            )
            # Add The parcelle size
            self.worksheet.write_formula(
                i,
                col + 2 + len(self.dfDroite.columns),
                f"={index + 1}*1.6",
                self.workbook.add_format({"align": "left"}),
            )

        # Hide columns
        self.worksheet.set_column("B:B", None, None, {"hidden": True})
        self.worksheet.set_column("AC:AC", None, None, {"hidden": True})

        # Set the layout options
        self.worksheet.set_pagebreak_view()
        self.worksheet.fit_to_pages(1, 1)
        self.worksheet.set_landscape()

    def picasso(self):
        """
        Mettre les coleurs dans le plan
        """
        wb = openpyxl.load_workbook(self.nomDuPlan + ".xlsx")
        sheet = wb["Plan"]
        # Create the PatternFill objects using the colors
        colorsLeft = [
            PatternFill(start_color=color, end_color=color, fill_type="solid")
            for color in self.colors[: len(self.gaucheIndex)]
        ]
        colorsRight = [
            PatternFill(start_color=color, end_color=color, fill_type="solid")
            for color in self.colors[len(self.gaucheIndex) :]
        ]
        gray_fill = PatternFill(
            start_color="C4C4C4", end_color="C4C4C4", fill_type="solid"
        )
        # Re-Create left and right matrix :D
        leftMatriceList = [self.essaisMatricesList[i] for i in self.gaucheIndex]
        rightMatriceList = [self.essaisMatricesList[i] for i in self.droiteIndex]

        # Color the essais at the left side
        col = self.ppInitialCol + 1
        row = self.ppInitialRow + 1
        for color, index in enumerate(self.gaucheIndex):
            blocs = self.essais[index].nBlocs
            modas = self.essais[index].nModalite
            for i in range(row, row + modas):
                for j in range(col, col + blocs * 2):
                    sheet.cell(row=i, column=j).fill = colorsLeft[color]
                for k in range(
                    self.initialCol + 1,
                    self.initialCol + 1 + len(self.dfGauche.columns),
                ):
                    sheet.cell(row=i, column=k).fill = colorsLeft[color]
            row = row + modas

        # Color the essais at the right side
        col = self.ppInitialCol + 1 + leftMatriceList[0].shape[1] * 2
        row = self.ppInitialRow + 1
        colIniBordure = (
            (self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 4) + 1 + 1 + 1
        )  # +1 pour laisser un espace et +1 pour la colonne d'énumeration
        for color, index in enumerate(self.droiteIndex):
            blocs = self.essais[index].nBlocs
            modas = self.essais[index].nModalite
            for i in range(row, row + modas):
                for j in range(col, col + blocs * 2):
                    sheet.cell(row=i, column=j).fill = colorsRight[color]
                for k in range(
                    colIniBordure, colIniBordure + len(self.dfDroite.columns)
                ):
                    sheet.cell(row=i, column=k).fill = colorsRight[color]
            row = row + modas

        # Mettre les bordures grises dans le plan
        planCol = self.ppInitialCol + 1
        for i in range(
            self.ppInitialRow + 1, self.ppInitialRow + len(self.dfGauche) + 1
        ):
            for j in range(planCol, planCol + 2):
                sheet.cell(row=i, column=j).fill = gray_fill

        planCol = (
            self.ppInitialCol
            + 1
            + ((leftMatriceList[0].shape[1] + rightMatriceList[0].shape[1]) * 2)
            - 2
        )
        for i in range(
            self.ppInitialRow + 1, len(self.dfDroite) + self.ppInitialRow + 1
        ):
            for j in range(planCol, planCol + 2):
                sheet.cell(row=i, column=j).fill = gray_fill

        wb.save(self.nomDuPlan + ".xlsx")
        print("Plan généré avec succès :D")


class Dessinateur2blocs(Dessinateur4blocs):
    def __init__(self, nomDuPlan, essais, essaisMatricesList):
        super().__init__(nomDuPlan, essais,
                         essaisMatricesList,
                         [], [])
    

    def planPrincipal(self):
        row = self.ppInitialRow
        col = self.ppInitialCol

        for matrice in self.essaisMatricesList:
            matrice = pd.DataFrame(matrice)
            matrice = matrice[matrice.columns[::-1]]
            for bloc in matrice.columns:
                self.worksheet.write_column(
                    row, col, list(matrice[bloc]), self.centerCellFormat
                )
                col += 2
            row += len(matrice)
            col = self.ppInitialCol

    
    def drawBordures(self):
        row = self.initialRow
        col = self.initialCol

        self.dfEssais = pd.concat(
            [essai.descriptionData for essai in self.essais]
        ).reset_index(drop=True)

        for colone in self.dfEssais.columns:
            self.worksheet.write(
                row - 1, col, colone, self.workbook.add_format({"bold": True})
            )
            self.worksheet.write_column(
                row, col, list(self.dfEssais[colone]), self.simpleCellFormat
            )
            col += 1
    
    def matriceCompteur(self):
        """
        Compte les microparcelles de chaque essai
        """
        matriceComptage = np.zeros(
            (
                sum(essai.nModalite for essai in self.essais),
                self.essais[0].nBlocs,
            )
        )

        matriceComptage[0, :] = range(1, matriceComptage.shape[1] + 1)
        for i in range(len(matriceComptage) - 1):
            matriceComptage[i + 1, :] = range(
                int(matriceComptage[i, -1] + 1),
                int(matriceComptage[i, -1]) + matriceComptage.shape[1] + 1,
            )

        for i, row in enumerate(matriceComptage):
            if i % 2 == 0:
                matriceComptage[i] = row[::-1]

        row = self.ppInitialRow
        col = self.ppInitialCol + 1

        self.matriceComptage = pd.DataFrame(matriceComptage)
        for colonne in self.matriceComptage.columns:
            self.worksheet.write_column(
                row, col, list(self.matriceComptage[colonne]), self.countCellFormat
            )
            col += 2

    def fisher(self):
        """
        Integre les formules mathematiques pour le calcul des poids des sacs
        """

        lenConfig = len(pd.read_excel(CONFIG_NAME))+1

        for i in range(len(self.dfEssais)):
            self.worksheet.write_formula(
                f"F{3+i}",
                f"=G{3+i}*H{3+i}*Calcul!$C$3*Calcul!$C$4/1000",
                self.simpleCellFormat,
            )

        # LookUp pour les densités grace au pave de la feuille Calcul

        for i in range(3, len(self.dfEssais) + 3):
            self.worksheet.write_formula(
                f"H{i}",
                f"=XLOOKUP(B{i},Calcul!E4:E10,Calcul!F4:F10,0)",
                self.simpleCellFormat,
            )
        
        # LookUp pour les PMGs qui s'actualisent en temps reél
        for i in range(3, len(self.dfEssais) + 3):
            self.worksheet.write_formula(
                f"G{i}",
                f"=XLOOKUP(A{i},'{os.getcwd()}\[config.xlsx]Feuil1'!B2:B{lenConfig},'{os.getcwd()}\[config.xlsx]Feuil1'!D2:D{lenConfig},0)",
                self.simpleCellFormat,
            )

    def etiquettes(self):
        """
        Construit l'onglet dediée à l'impression des etiquettes
        """

        startPoint = 0
        etiquettes = pd.DataFrame(columns=["Position", "Parcelle", "Essai"])
        for essai, matrice in zip(self.essais, self.essaisMatricesList):
            position = (
                self.matriceComptage[startPoint : startPoint + essai.nModalite]
                .to_numpy()
                .flatten()
            )
            matrice = pd.DataFrame(matrice)
            matrice = matrice[matrice.columns[::-1]]
            parcelle = matrice.to_numpy().flatten()
            newEtiquetets = pd.DataFrame(
                {"Position": position, "Parcelle": parcelle, "Essai": essai.nom}
            )
            newEtiquetets["Modalité"] = (
                newEtiquetets["Parcelle"] - newEtiquetets["Parcelle"] // 100 * 100
            )
            newEtiquetets = pd.merge(
                newEtiquetets,
                essai.descriptionData.loc[:, ["VARIETES", "N"]],
                left_on="Modalité",
                right_on="N",
            )
            etiquettes = pd.concat([etiquettes, newEtiquetets])
            startPoint += essai.nModalite

        etiquettes = etiquettes.loc[
            :, ["Position", "Parcelle", "Essai", "VARIETES", "Modalité"]
        ]
        etiquettes["Plan"] = self.nomAffichage
        etiquettes = etiquettes.rename(columns={"VARIETES": "Variété"})
        etiquettes = etiquettes.sort_values(by="Position")
        etiquettes = etiquettes.reset_index(drop=True)

        worksheetEtiquettes = self.workbook.add_worksheet("Etiquettes")
        for colonne in etiquettes.columns:
            worksheetEtiquettes.write_column(
                1,
                list(etiquettes.columns).index(colonne),
                list(etiquettes[colonne]),
                self.simpleCellFormat,
            )
        worksheetEtiquettes.write_row(
            0, 0, list(etiquettes.columns), self.countCellFormat
        )

        worksheetEtiquettes.write(
            0, len(etiquettes.columns), "P.SAC", self.countCellFormat
        )

        for i in range(2, len(etiquettes) + 2):
            worksheetEtiquettes.write_formula(
                f"G{i}",
                f"=XLOOKUP(C{i}&D{i},Plan!D{self.ppInitialRow+1}:D{self.ppInitialRow + len(self.dfEssais)}&Plan!A{self.ppInitialRow+1}:A{self.ppInitialRow + len(self.dfEssais)},Plan!F{self.ppInitialRow+1}:F{self.ppInitialRow + len(self.dfEssais)},0)",
                self.simpleCellFormat,
            )
        worksheetEtiquettes.autofit()
    
    def maquillage(self):
        """
        Regle les proportions des celules et effectue la mise en page finale
        """
        self.worksheet.autofit()
        # Regler les proportions des colonnes
        for i in range(
            self.ppInitialCol,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 2,
            2,
        ):
            self.worksheet.set_column(i, i, 12)

        # Mettre le titre dans le plan principal
        self.worksheet.merge_range(
            self.ppInitialRow - 2,
            self.ppInitialCol,
            self.ppInitialRow - 2,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 2 - 1,
            self.nomAffichage,
            self.workbook.add_format({"bold": "true", "align": "center", "border": 1}),
        )

        # Fusionner les cellules des bordures
        for i in range(
            self.ppInitialCol,
            self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 2,
            2,
        ):
            self.worksheet.merge_range(
                self.ppInitialRow - 1,
                i,
                self.ppInitialRow - 1,
                i + 1,
                "",
                self.workbook.add_format(
                    {"bold": "true", "align": "center", "border": 1}
                ),
            )
            self.worksheet.merge_range(
                self.ppInitialRow + len(self.dfEssais),
                i,
                self.ppInitialRow + len(self.dfEssais),
                i + 1,
                "",
                self.workbook.add_format(
                    {"bold": "true", "align": "center", "border": 1}
                ),
            )

        # Add the counter at right of the plan
        col = self.ppInitialCol + self.essaisMatricesList[0].shape[1] * 2
        row = self.ppInitialRow - 1
        for index, i in enumerate(range(row, row + len(self.dfEssais) + 2)):
            self.worksheet.write(
                i, col,
                index + 1,
                self.workbook.add_format({"align": "left"})
            )
            # Add The parcelle size
            self.worksheet.write_formula(
                i,
                col + 2,
                f"={index + 1}*1.6",
                self.workbook.add_format({"align": "left"}),
            )

        # Hide columns
        self.worksheet.set_column("B:B", None, None, {"hidden": True})

        # Set the layout options
        self.worksheet.set_pagebreak_view()
        self.worksheet.fit_to_pages(1, 1)
        self.worksheet.set_landscape()
    

    def picasso(self):
        """
        Mettre les coleurs dans le plan
        """
        wb = openpyxl.load_workbook(self.nomDuPlan + ".xlsx")
        sheet = wb["Plan"]
        # Create a PatternFill
        colorsFill = [
            PatternFill(start_color=color, end_color=color, fill_type="solid")
            for color in self.colors
        ]
        gray_fill = PatternFill(
            start_color="C4C4C4", end_color="C4C4C4", fill_type="solid"
        )

        # Color the essais
        col = self.ppInitialCol + 1
        row = self.ppInitialRow + 1
        for color, index in enumerate(range(len(self.essais))):
            blocs = self.essais[index].nBlocs
            modas = self.essais[index].nModalite
            for i in range(row, row + modas):
                for j in range(col, col + blocs * 2):
                    sheet.cell(row=i, column=j).fill = colorsFill[color]
                for k in range(
                    self.initialCol + 1,
                    self.initialCol + 1 + len(self.dfEssais.columns),
                ):
                    sheet.cell(row=i, column=k).fill = colorsFill[color]
            row = row + modas

        # Color the essais at the right bordure
        planCol = self.ppInitialCol - 1 + self.essaisMatricesList[0].shape[1] * 2
        for i in range(
            self.ppInitialRow + 1, len(self.dfEssais) + self.ppInitialRow + 1
        ):
            for j in range(planCol, planCol + 2):
                sheet.cell(row=i, column=j).fill = gray_fill

        wb.save(self.nomDuPlan + ".xlsx")
        print("Plan généré avec succès :D")


def lineToMicroparcelle(row):
    """
    A fonction to use into the pandas .apply() method to create a microparcelle object
    """
    nom = int(row["Position"] / 100) * 100 + row["N° de traitement"]
    position = row["Position"]
    traitement = row["N° de traitement"]
    return Microparcelle(nom, position, traitement)


def creerEssais(srcPath, pathListe):
    essais = [
        Essai(path[:-5], 0, srcPath + path, CONFIG_NAME)
        for path in pathListe
    ]
    for essai in essais:
        essai.createEssai()
        essai.create_matrice()
    return essais


def orchesterPlan(listeEssais):
    """
    Orchestre le plan en balançant les essais sur la base de sa longueur.
    Utilise la force brute pour trouver la meilleure combinaison.
    """

    n = len(listeEssais)
    total_sum = sum(listeEssais)
    
    best_diff = float('inf')  # Initialize with a very large number
    best_partition = None

    indices = list(range(2, n))  # List of indices, excluding the first two elements

    # Try all possible subsets of the remaining indices
    for i in range(len(indices) + 1):  # For each possible size of the subset
        for subset_indices in combinations(indices, i):
            subset1_indices = [0] + list(subset_indices)  # Ensure index 0 is in subset1
            subset2_indices = [1] + list(set(indices) - set(subset_indices))  # Ensure index 1 is in subset2

            subset1_sum = sum(listeEssais[idx] for idx in subset1_indices)
            subset2_sum = sum(listeEssais[idx] for idx in subset2_indices)
            diff = abs(subset1_sum - subset2_sum)  # Calculate the difference in sums

            # Update the best partition if the difference is smaller
            if diff < best_diff:
                best_diff = diff
                best_partition = (subset1_indices, subset2_indices)

    subset1_indices, subset2_indices = best_partition
    subset1_indices.sort()
    subset2_indices.sort()

    if sum(listeEssais[idx] for idx in subset2_indices) > sum(listeEssais[idx] for idx in subset1_indices) :
        return subset2_indices, subset1_indices
    else:
        return subset1_indices, subset2_indices


def backendWrapper4Blocs(srcPath):
    """
    Une fonction qui relie toutes les actions nécessaires
    pour la creation et l'organisation des essais dans le cas d'un site à 4 blocs.
    """
    # Arranger et modifier la liste de fichiers pour garantir la position des essais Gamme et bt2
    os.makedirs(srcPath + "results", exist_ok=True)
    docs = os.listdir(srcPath)
    docs = [doc for doc in docs if doc.endswith(".xlsx")]
    docs.remove("Gamme.xlsx")
    docs.remove("bt2.xlsx")
    docs.insert(0, "bt2.xlsx")
    docs.insert(1, "Gamme.xlsx")

    essais = creerEssais(srcPath, docs)
    essaisMatricesList = [essai.matrice for essai in essais]

    gaucheIndex, droiteIndex = orchesterPlan(
        [len(matrice) for matrice in essaisMatricesList]
    )

    return essais, essaisMatricesList, gaucheIndex, droiteIndex

def backendWrapper2Blocs(srcPath):
    """
    Une fonction qui relie toutes les actions nécessaires
    pour la creation et l'organisation des essais dans le cas d'un site à 2 blocs.
    """
    # Creer un dossier pour les fichiers de sortie
    os.makedirs(srcPath + "results", exist_ok=True)
    docs = os.listdir(srcPath)
    docs = [doc for doc in docs if doc.endswith(".xlsx")]
    docs.remove("Gamme.xlsx")
    docs.insert(0, "Gamme.xlsx")
    essais = creerEssais(srcPath, docs)
    essaisMatricesList = [essai.matrice for essai in essais]

    return essais, essaisMatricesList
